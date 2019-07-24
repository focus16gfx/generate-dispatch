from openpyxl import Workbook, load_workbook
import openpyxl
from openpyxl.styles.borders import BORDER_THIN, BORDER_THICK
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, colors
import datetime
import re
import os
import sys
import csv
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

pickupdate_index = 3
deliverydate_index = 5
print('>>Initialising...')
wd = os.getcwd()
# wd = '\\\ATL09FPS01\Accord-Folders\sschmidt\Desktop\Dispatch_script\Dispatch_script'



redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
yellowFill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
greenFill = PatternFill(start_color='FF00cc00', end_color='FF006600', fill_type='solid')
blueFill = PatternFill(start_color='FF0099ff', end_color='FF0099ff', fill_type='solid')
lightblueFill = PatternFill(start_color='FFC5D9F1', end_color='FFC5D9F1', fill_type='solid')
blackFill = PatternFill(start_color=colors.BLACK, end_color=colors.BLACK, fill_type='solid')
lightPurpleFill = PatternFill(start_color='FFB1A0C7', end_color='FFB1A0C7', fill_type='solid')
inactiveGreyFill = PatternFill(start_color='FF808080', end_color='FF808080', fill_type='solid')
#B1A0C7   808080

# --------------------------------------------------------------------------------


ASSGNFill = PatternFill(start_color='FFff0000', end_color='FFff0000', fill_type='solid')
DISPFill = PatternFill(start_color='FFffa500', end_color='FFffa500', fill_type='solid')
ATPICKFill = PatternFill(start_color='FFffff00', end_color='FFffff00', fill_type='solid')
PICKDFill = PatternFill(start_color='FF008000', end_color='FF008000', fill_type='solid')
BKRPICKDFill = PatternFill(start_color='FFff00ff', end_color='FFff00ff', fill_type='solid')
BORDERFill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
ENRTEFill = PatternFill(start_color='FF9e379f', end_color='FF800080', fill_type='solid')
CARDEDFill = PatternFill(start_color='FF60a1f0', end_color='FF800080', fill_type='solid')
ATCONSFill = PatternFill(start_color=colors.BLUE, end_color=colors.BLUE, fill_type='solid')
SP4DELFill = PatternFill(start_color='FF4b0082', end_color='FF4b0082', fill_type='solid')
SP4OBFill = PatternFill(start_color='FFee82ee', end_color='FFee82ee', fill_type='solid')
SPTLDFill = PatternFill(start_color='FF00ffff', end_color='FF00ffff', fill_type='solid')
thin_border = Border(bottom=Side(border_style=BORDER_THIN, color='00000000', ))
thin_allborder = Border(bottom=Side(border_style=BORDER_THIN, color='00000000'),
                        top=Side(border_style=BORDER_THIN, color='00000000'),
                        left=Side(border_style=BORDER_THIN, color='00000000'),
                        right=Side(border_style=BORDER_THIN, color='00000000'))
thick_border = Border(bottom=Side(border_style=BORDER_THICK, color='00000000'),
                      top=Side(border_style=BORDER_THICK, color='00000000'),
                      left=Side(border_style=BORDER_THICK, color='00000000'),
                      right=Side(border_style=BORDER_THICK, color='00000000'))

'''create font objects for different text fields'''
ft = Font(bold=True, size=15)
ft_small = Font(bold=True)
ft_white = Font(color=colors.WHITE)
ft_grey = Font(color='FF808080')
title_ft = Font(bold=True, size=25)
title_2ft = Font(bold=True, size=20)
ft_bld_black_14 = Font(color=colors.BLACK, bold=True, size=12)


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)
    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment
    rows = ws[cell_range]
    if font:
        for row in rows:
            for cell in row:
                cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom
    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def create_orig_dest_sheet(og_wb, origin='', destination='', origin_everywhere=False, destination_everywhere=False):
    og_ws_main = og_wb.active
    if origin_everywhere:
        og_ws_O2D = og_wb.create_sheet(f'{destination} Inbound')
    if destination_everywhere:
        og_ws_O2D = og_wb.create_sheet(f'{origin} Outbound')

    if not origin_everywhere and not destination_everywhere:
        og_ws_O2D = og_wb.create_sheet(f'{origin} to {destination}')

    '''write headers'''
    for i in range(12):
        og_ws_O2D.cell(1, i + 1).value = og_ws_main.cell(1, i + 1).value
        og_ws_O2D.cell(1, i + 1).alignment = Alignment(horizontal='center')


    '''append rows to new origin to destination sheet'''
    O2D_rows = []
    prev_trip = ''
    current_trip = ''
    for row in list(og_ws_main.rows)[2:]:
        if row[2].value:
            origin_pr = row[2].value[-3:].replace(' ', '')
        else:
            origin_pr = ''
        if row[4].value:
            destination_pr = row[4].value[-3:].replace(' ', '')
        else:
            destination_pr = ''
        origin_full = row[2].value
        destination_full = row[4].value
        # current_trip = row[1].value
        # if str(current_trip) == '721467':
        #     print()

        ###SUPRESS BROKERED LOADS USING K COLUMN. If IT'S EMPTY...NOT BROKERED
        if len(row[10].value) > 2:
            continue

        if origin_everywhere:
            if destination_pr == destination:
                """supress local loads..ex: AB-->AB, BC--> etc"""
                if origin_pr == destination:
                    continue

                """MINUS 1 DAY IF CONDITION SATISFIES"""
                FB_string = row[11].value
                if origin_full in ['HOPE BC', 'RICHMOND, BC',
                                   'CHILLIWACK, BC'] and destination_pr == 'AB' and FB_string == 'WAL-MART CANADA C/O ABM':
                    delta_oneday = datetime.timedelta(days=1)
                    deliverby_string = row[5].value
                    fmts = ['%m/%d/%Y %H:%M:%S', '%m/%d/%Y']
                    for fmt in fmts:
                        try:
                            dt_dlvry = datetime.datetime.strptime(deliverby_string, fmt)
                            dt_dlvry = dt_dlvry - delta_oneday
                            deliverby_string = dt_dlvry.strftime(fmt).lstrip("0")
                            row[5].value = deliverby_string
                            break
                        except:
                            pass

                """skip rows with same trip as last trip"""
                current_trip = row[1].value
                if current_trip == prev_trip:
                    O2D_rows.pop(-1)
                O2D_rows.append(row)
                prev_trip = current_trip


        if destination_everywhere:
            if origin_pr == origin:
                """supress local loads..ex: AB-->AB, BC--> etc"""
                if destination_pr == origin:
                    continue

                """MINUS 1 DAY IF CONDITION SATISFIES"""
                FB_string = row[11].value
                if origin_full in ['HOPE BC', 'RICHMOND, BC',
                                   'CHILLIWACK, BC'] and destination_pr == 'AB' and FB_string == 'WAL-MART CANADA C/O ABM':
                    delta_oneday = datetime.timedelta(days=1)
                    deliverby_string = row[5].value
                    fmts = ['%m/%d/%Y %H:%M:%S', '%m/%d/%Y']
                    for fmt in fmts:
                        try:
                            dt_dlvry = datetime.datetime.strptime(deliverby_string, fmt)
                            dt_dlvry = dt_dlvry - delta_oneday
                            deliverby_string = dt_dlvry.strftime(fmt).lstrip("0")
                            row[5].value = deliverby_string
                            break
                        except:
                            pass

                """skip rows with same trip as last trip"""
                current_trip = row[1].value
                if current_trip == prev_trip:
                    O2D_rows.pop(-1)
                O2D_rows.append(row)
                prev_trip = current_trip

            """Special case. WA-AB records in BC-ALL AKA BC OUTBOUND"""
            if origin == 'BC' and ((origin_pr == 'WA' and destination_pr == 'AB') or ((origin_full == 'Wayfair Perris, CA' or origin_full.lower() == 'perris, ca') and destination_full.lower() == 'genelle, bc')):
                """skip rows with same trip as last trip"""
                current_trip = row[1].value
                if current_trip == prev_trip:
                    O2D_rows.pop(-1)
                O2D_rows.append(row)
                prev_trip = current_trip

        if not origin_everywhere and not destination_everywhere:
            if origin_pr == origin and destination_pr == destination:

                """MINUS 1 DAY IF CONDITION SATISFIES"""
                FB_string = row[11].value
                if origin_full in ['HOPE BC', 'RICHMOND, BC',
                                   'CHILLIWACK, BC'] and destination_pr == 'AB' and FB_string == 'WAL-MART CANADA C/O ABM':
                    delta_oneday = datetime.timedelta(days=1)
                    deliverby_string = row[5].value
                    fmts = ['%m/%d/%Y %H:%M:%S', '%m/%d/%Y']
                    for fmt in fmts:
                        try:
                            dt_dlvry = datetime.datetime.strptime(deliverby_string, fmt)
                            dt_dlvry = dt_dlvry - delta_oneday
                            deliverby_string = dt_dlvry.strftime(fmt).lstrip("0")
                            row[5].value = deliverby_string
                            break
                        except:
                            pass

                """skip rows with same trip as last trip"""
                current_trip = row[1].value
                if current_trip == prev_trip:
                    O2D_rows.pop(-1)
                O2D_rows.append(row)
                prev_trip = current_trip


    """write row data to sheet"""
    row_pointer = 3
    for row in O2D_rows:
        for column, cellu in enumerate(row, start=1):
            og_ws_O2D.cell(row_pointer, column).value = cellu.value
            og_ws_O2D.cell(row_pointer, column).alignment = Alignment(horizontal='center')
        row_pointer += 1

    return og_ws_O2D


def csv_to_xlsx(og_filename):
    """take filename as input and convert csv to xlsx"""
    # og_filename = input('**Report csv file name: ')
    wb = openpyxl.Workbook()
    ws = wb.active
    with open(f'{wd}\\{og_filename}.csv') as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)
    wb.save(f'{wd}\\{og_filename}.xlsx')

    og_wb = load_workbook(f'{wd}\\{og_filename}.xlsx')
    og_ws_main = og_wb.active

    '''delete all nulls and center align main sheet'''
    for row in list(og_ws_main.rows)[2:]:
        for column, cellu in enumerate(row, start=1):
            cellu.alignment = Alignment(horizontal='center')
            if cellu.value.replace(' ', '') == '<null>':
                cellu.value = ''

    # og_ws_BC2AB = create_orig_dest_sheet(og_wb, 'BC', 'AB')
    # og_ws_AB2BC = create_orig_dest_sheet(og_wb, 'AB', 'BC')
    # og_ws_AB2BC = create_orig_dest_sheet(og_wb, 'CA', 'BC')
    # og_ws_WA2BC = create_orig_dest_sheet(og_wb, 'WA', 'BC')
    og_ws_Everywhere2BC = create_orig_dest_sheet(og_wb, destination='BC', origin_everywhere=True)
    og_ws_BC2Everywhere = create_orig_dest_sheet(og_wb, destination_everywhere=True, origin='BC')
    og_ws_Everywhere2BC = create_orig_dest_sheet(og_wb, destination='AB', origin_everywhere=True)
    og_ws_AB2Everywhere = create_orig_dest_sheet(og_wb, destination_everywhere=True, origin='AB')


    '''get all sheet names'''
    sheets = og_wb.sheetnames

    '''adjust column width'''
    for sheet_title in sheets:
        ws = og_wb[sheet_title]
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 13
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 30
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 40


    og_wb.save(f'{wd}\\{og_filename}.xlsx')


'''parse unique dates from Delivery dates'''


def get_dates(group_by=deliverydate_index):
    global dates
    pt3 = re.compile(r'^\d{1,2}/\d{1,2}/\d{4}')
    for row in rows_list:
        date_extracted = re.match(pt3, str(row[group_by].value)).group(0)
        if date_extracted not in dates:
            dates.append(date_extracted)


def write_headers(row_num, fontyy=ft):
    """write headers to a given row in new sheet"""
    for coord in dict_headers:
        ws[coord + str(row_num)] = dict_headers.get(coord)
        ws[coord + str(row_num)].font = fontyy
        ws[coord + str(row_num)].fill = lightblueFill
        ws[coord + str(row_num)].alignment = Alignment(horizontal='center')
    style_range(ws, f'A{row_num}:J{row_num}', border=thin_allborder)


def each_date(date, sheet_name, group_by=deliverydate_index, sort_by=pickupdate_index):
    global lol
    queue_list = []
    queue_list = []
    '''Add all rows in new rowlist with this date'''
    for row_q1 in rows_list:
        dt6 = datetime.datetime.strptime(date, '%m/%d/%Y')
        match_date = dt6.strftime('%A, %B %d, %Y')
        delta_one_day = datetime.timedelta(days=1)

        if match_date in str(row_q1[group_by].value):
            queue_list.append(row_q1)

    if old_filename:
        '''------------old------------'''
        old_queue_list = []
        '''Add all rows in new rowlist with this date'''
        for row_q1 in old_row_list:
            dt6 = datetime.datetime.strptime(date, '%m/%d/%Y')
            match_date = dt6.strftime('%A, %B %d, %Y')
            if match_date in str(row_q1[group_by].value):
                old_queue_list.append(row_q1)

        '''sort new rowlist by pickup'''
        old_queue_list.sort(key=lambda x: datetime.datetime.strptime(str(x[sort_by].value), '%A, %B %d, %Y @ %H:%M'))

        '''manipulation op'''
        for new_row in queue_list:
            new_trip = new_row[1].value
            for old_row in old_queue_list:
                if old_row[1].value == new_trip:
                    n_pu = new_row[6].value
                    n_trailer = new_row[7].value
                    n_status = new_row[8].value.strip()
                    n_probill = new_row[0].value
                    n_notes = new_row[9].value
                    o_pu = old_row[6].value
                    o_trailer = old_row[7].value
                    o_status = old_row[8].value.strip()
                    o_probill = old_row[0].value
                    o_notes = old_row[9].value
                    if not n_pu:
                        n_pu = o_pu
                        new_row[6].value = n_pu
                    if not n_trailer:
                        n_trailer = o_trailer
                        new_row[7].value = n_trailer
                    if not n_probill:
                        n_probill = o_probill
                        new_row[0].value = n_probill

                    n_notes = o_notes
                    new_row[9].value = n_notes
                    # ##add all notes from old document
                    # try:
                    #     new_row[9] = old_row[9].value
                    # except TypeError:
                    #     new_row[9] = ''

                    prior_list = ['ASSGN', 'DISP', 'ATPICK', 'PICKD', 'BKRPICKD', 'BORDER', 'ENRTE', 'ATCONS', 'SP4DEL', 'SP4OB', 'CARDED', 'SPTLD', 'DELVD']
                    status_priority = {}
                    for priority_index, status in enumerate(prior_list, start=1):
                        status_priority[status] = priority_index

                                    #'CARDED '


                    if o_status == 'BROKER':
                        n_status = o_status
                        new_row[8].value = n_status
                    elif status_priority[o_status] > status_priority[n_status]:
                        n_status = o_status
                        new_row[8].value = n_status
    if old_filename:
        for old_row in old_queue_list:
            lol = []
            '''manual update'''
            if str(old_row[1].value)[-1].lower() == 'm':

                for i in range(10):
                    cell = og_sheet['AH8']
                    lol.append(cell)

                lol[1] = old_row[1]
                lol[3] = old_row[3]
                lol[2] = old_row[2]
                lol[5] = old_row[5]
                lol[4] = old_row[4]
                lol[0] = old_row[0]
                lol[6] = old_row[6]
                lol[7] = old_row[7]
                lol[8] = old_row[8]
                queue_list.append(lol)

    '''sort new rowlist by  pickup'''

    queue_list.sort(key=lambda x: datetime.datetime.strptime(str(x[sort_by].value), '%A, %B %d, %Y @ %H:%M'))

    global current_row
    #Monday, July 22, 2019
    '''add heading each day'''
    dt = datetime.datetime.strptime(date, '%m/%d/%Y')
    ws[f'A{current_row}'] = dt.strftime('%A, %B %d, %Y')
    ws.merge_cells(f'A{current_row}:J{current_row}')
    ws[f'A{current_row}'].font = ft
    current_row += 1
    ws[f'A{current_row}'] = 'Date Notes:'
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'].font = ft
    ws[f'A{current_row}'].fill = lightblueFill
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    ws[f'C{current_row}'] = ''
    ws[f'C{current_row}'].fill = lightPurpleFill
    # ws[f'C{current_row}'].alignment = Alignment(horizontal='center')
    ws[f'C{current_row}'].font = ft_bld_black_14
    ws.merge_cells(f'C{current_row}:J{current_row}')
    style_range(ws, f'A{current_row}:J{current_row}', border=thin_allborder)

    current_row += 1

    write_headers(current_row, ft_small)

    '''write sorted list of rows to excel'''
    print('PICKUP DATES...')

    style_range(ws, f'A{current_row}:J{current_row + len(queue_list)}', border=thin_allborder)


    for row_q in queue_list:
        current_row += 1
        """Color records in grey if Origin,Dest is not AB,BC"""
        origin_full = row_q[2].value
        destination_full = row_q[4].value
        options_o_d = ['AB', 'BC']
        if row_q[2].value:
            origin_pr = row_q[2].value[-3:].replace(' ', '')
        else:
            origin_pr = ''
        if row_q[4].value:
            destination_pr = row_q[4].value[-3:].replace(' ', '')
        else:
            destination_pr = ''
        if origin_pr not in options_o_d or destination_pr not in options_o_d:
            if sheet_name == 'BC Outbound' and ((origin_pr == 'WA' and destination_pr == 'AB') or ((origin_full == 'Wayfair Perris, CA' or origin_full.lower() == 'perris, ca') and destination_full.lower() == 'genelle, bc')):
                pass
            else:
                style_range(ws, f'A{current_row}:I{current_row}', font=ft_grey)

        ws[f'D{current_row}'] = row_q[3].value
        ws[f'C{current_row}'] = row_q[2].value
        ws[f'E{current_row}'] = row_q[4].value
        ws[f'F{current_row}'] = row_q[5].value
        ws[f'B{current_row}'] = row_q[1].value
        ws[f'A{current_row}'] = row_q[0].value

        ws[f'G{current_row}'] = row_q[6].value
        if row_q[6].value is None:
            ws[f'G{current_row}'].fill = yellowFill

        ws[f'H{current_row}'] = row_q[7].value
        if row_q[7].value is None:
            ws[f'H{current_row}'].fill = yellowFill

        ws[f'I{current_row}'] = row_q[8].value

        ws[f'J{current_row}'] = row_q[9].value

        status = row_q[8].value

        '''color fill'''
        if status == 'ASSGN':
            ws[f'I{current_row}'].fill = ASSGNFill
        elif status == 'DISP':
            ws[f'I{current_row}'].fill = DISPFill
        elif status == 'ATPICK':
            ws[f'I{current_row}'].fill = ATPICKFill
        elif status == 'PICKD':
            ws[f'I{current_row}'].fill = PICKDFill
        elif status == 'BKRPICKD':
            ws[f'I{current_row}'].fill = BKRPICKDFill
        elif status == 'BORDER':
            ws[f'I{current_row}'].fill = BORDERFill
            ws[f'I{current_row}'].font = ft_white


        elif status == 'CARDED':
            ws[f'I{current_row}'].fill = CARDEDFill

        elif status == 'ENRTE':
            # ws[f'I{current_row}'].fill = ENRTEFill
            style_range(ws, f'A{current_row}:I{current_row}', fill=ENRTEFill)

        elif status == 'ATCONS':
            ws[f'I{current_row}'].fill = ATCONSFill
            ws[f'I{current_row}'].font = ft_white
        elif status == 'SP4DEL':
            ws[f'I{current_row}'].fill = SP4DELFill
            ws[f'I{current_row}'].font = ft_white
        elif status == 'SP4OB':
            ws[f'I{current_row}'].fill = SP4OBFill
        elif status == 'SPTLD':
            ws[f'I{current_row}'].fill = SPTLDFill

        elif status == 'BROKER':
            style_range(ws, f'A{current_row}:I{current_row}', fill=ENRTEFill)

        elif status == 'DELVD':
            style_range(ws, f'A{current_row}:I{current_row}', fill=blackFill)
            columns = ws[f'A{current_row}:I{current_row}']
            for cell in columns[0]:
                cell.font = ft_white





        '''center align'''
        ws[f'D{current_row}'].alignment = Alignment(horizontal='center')
        ws[f'C{current_row}'].alignment = Alignment(horizontal='center')
        ws[f'E{current_row}'].alignment = Alignment(horizontal='center')
        ws[f'F{current_row}'].alignment = Alignment(horizontal='center')
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center')
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
        ws[f'G{current_row}'].alignment = Alignment(horizontal='center')
        ws[f'H{current_row}'].alignment = Alignment(horizontal='center')
        ws[f'I{current_row}'].alignment = Alignment(horizontal='center')


        print(f'-{row_q[5].value}')

print('>>Script started!')

'''take filename as input and convert csv to xlsx'''
og_filename = None
for file in os.listdir(path=wd):
    if 'EXPORT' in file:
        og_filename = file[:-4]
        print(f'>>Found csv Report file, {file}!!')
        break
if not og_filename:
    og_filename = input('**Report csv file name: ')
# og_filename = 'FBSTATUS'
csv_to_xlsx(og_filename)





"""Open xlsx converted original sheet"""
og_wb = openpyxl.load_workbook(f'{wd}\\{og_filename}.xlsx')
# og_sheet = og_wb.active

sheets = og_wb.sheetnames
for sheet in sheets[1:]:

    old_filename = None
    for file in os.listdir(path=wd):
        if sheet in file:
            old_filename = file[:-5]
            print(f'>>Found {sheet} old dispatch plan, {file}')
            break

    if not old_filename:
        old_filename = input(f'**{sheet} Old Dispatch plan file name: ')


    if old_filename == '':
        old_filename = None

    old_ws = ''

    if old_filename:
        old_wb = load_workbook(filename=wd + '\\' + old_filename + '.xlsx')
        old_ws = old_wb.active

    '''instantiate new worksheet to write data into'''
    wb = Workbook()
    ws = wb.active

    '''headers dictionary'''
    dict_headers = {'A': 'PROBILL',  # 0
                    'B': 'TRIP',  # 1
                    'C': 'ORIGIN',  # 2
                    'D': 'PICKUP BY',  # 3
                    'E': 'DESTINATION',  # 4
                    'F': 'DELIVER BY',  # 5
                    'G': 'P/U',  # 6
                    'H': 'TRAILER',  # 7
                    'I': 'STATUS',  #8
                    'J': 'NOTES'}  # 9

    old_dict_headers = {'J': 'TRIP',  # 9
                        'K': 'PICKUP BY',  # 10
                        'L': 'ORIGIN',  # 11
                        'Q': 'DELIVER BY',  # 16
                        'T': 'DESTINATION',  # 19
                        'U': 'PROBILL',  # 20
                        'Y': 'P/U',  # 24
                        'Z': 'TRAILER',  # 25
                        'AC': 'STATUS'}  # 28

    '''setting default width for columns'''
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 20

    """name"""
    sheet_title = sheet
    origin = sheet_title.split(' ')[0]
    destination = sheet_title.split(' ')[-1]
    sheet_title_manipped = sheet_title
    if origin == 'Everywhere':
        sheet_title_manipped = f'{destination} Inbound'
    elif destination == 'Everywhere':
        sheet_title_manipped = f'{origin} Outbound'


    og_sheet = og_wb[sheet_title]
    '''merge and format cells'''
    ws.merge_cells('A1:J1')
    ws['A1'].value = f'{sheet_title_manipped} PLANNER'
    ws['A1'].font = title_ft
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:J2')
    now_date = datetime.datetime.now()
    ws['A2'].value = now_date.strftime('%A, %B %d, %Y %H:%M')
    ws['A2'].font = title_2ft
    ws['A2'].alignment = Alignment(horizontal='center')


    ##delete FBSTATUS Column
    for cell in og_sheet['J']:
        cell.value = ''

    # make a regex pattern which searchs if data is present in column number 10
    pattern1 = re.compile(r'^\d{3,10}')
    rows_list = []
    '''collect the rows which contain data in rows_list'''
    for row in og_sheet.rows:
        value1 = str(row[1].value)
        if re.search(pattern1, value1):
            '''remove probills start with 'V' if applicable'''
            if is_import_v:
                rows_list.append(row)
            else:
                if not str(row[0].value).lower().startswith('v'):
                    rows_list.append(row)


    old_row_list = []
    '''get data from old sheet'''
    for row in old_ws:
        value1 = str(row[1].value)
        if re.search(pattern1, value1):
            old_row_list.append(row)

    dates = []
    if sheet_title == 'AB to BC':
        group_by = pickupdate_index
        sort_by = deliverydate_index
        get_dates(group_by)
    elif sheet_title == 'BC Inbound':
        group_by = pickupdate_index
        sort_by = deliverydate_index
        get_dates(group_by)
    elif sheet_title == 'BC Outbound':
        group_by = deliverydate_index
        sort_by = pickupdate_index
        get_dates(group_by)
    elif sheet_title == 'AB Inbound':
        group_by = pickupdate_index
        sort_by = deliverydate_index
        get_dates(group_by)
    elif sheet_title == 'AB Outbound':
        group_by = deliverydate_index
        sort_by = pickupdate_index
        get_dates(group_by)
    else:
        get_dates()



    dates.sort(key=lambda x: datetime.datetime.strptime(str(x), '%m/%d/%Y'))

    current_row = 4
    date_without_time_regex = re.compile(r'\d{4}$')
    '''add HH:MM to date if not present'''
    for row in rows_list:
        pick_dt = row[pickupdate_index].value
        del_dt = row[deliverydate_index].value
        if re.search(date_without_time_regex, pick_dt):
            row[pickupdate_index].value = str(row[pickupdate_index].value) + ' 00:00:00'
        if re.search(date_without_time_regex, del_dt):
            row[deliverydate_index].value = str(row[deliverydate_index].value) + ' 00:00:00'

    '''date necessary format'''
    for row in rows_list:
        dt3 = datetime.datetime.strptime(str(row[pickupdate_index].value), '%m/%d/%Y %H:%M:%S')
        row[pickupdate_index].value = dt3.strftime('%A, %B %d, %Y @ %H:%M')
        dt3 = datetime.datetime.strptime(str(row[deliverydate_index].value), '%m/%d/%Y %H:%M:%S')
        row[deliverydate_index].value = dt3.strftime('%A, %B %d, %Y @ %H:%M')

    print('----------------------------------------------------')

    for date in dates:
        print(f'>>Writing data for {date}')
        if sheet_title == 'AB to BC':
            group_by = pickupdate_index
            sort_by = deliverydate_index
            each_date(date, group_by=group_by, sort_by=sort_by, sheet_name=sheet_title)
        elif sheet_title == 'BC Inbound':
            group_by = pickupdate_index
            sort_by = deliverydate_index
            each_date(date, group_by=group_by, sort_by=sort_by, sheet_name=sheet_title)
        elif sheet_title == 'BC Outbound':
            group_by = deliverydate_index
            sort_by = pickupdate_index
            each_date(date, group_by=group_by, sort_by=sort_by, sheet_name=sheet_title)
        elif sheet_title == 'AB Inbound':
            group_by = pickupdate_index
            sort_by = deliverydate_index
            each_date(date, group_by=group_by, sort_by=sort_by, sheet_name=sheet_title)
        elif sheet_title == 'AB Outbound':
            group_by = deliverydate_index
            sort_by = pickupdate_index
            each_date(date, group_by=group_by, sort_by=sort_by, sheet_name=sheet_title)
        else:
            each_date(date, sheet_name=sheet_title)
        current_row += 3

    # ws.sheet_view.showGridLines = False

    name_ext = now_date.strftime('%A, %B %d, %Y %H:%M')
    # Tuesday, July 2, 2019 21:43

    dt_a1 = datetime.datetime.strptime(name_ext, '%A, %B %d, %Y %H:%M')
    name_ext = dt_a1.strftime('%A_%B_%d_%Y t_%H%M')


    name_ext = f'{sheet_title_manipped}_{name_ext}'
    out_fn = f'{name_ext}'
    folder = 'Output\\'
    try:
        os.mkdir(f'{wd}\\{folder[:-1]}')
    except FileExistsError:
        pass

    wb.save(f'{wd}\\{folder}{out_fn}.xlsx')

    print(f'>>SUCCESSFULLY generated a new dispatch plan, file saved in "{folder}{out_fn}.xlsx"')

    print('=============================================================================================================')
